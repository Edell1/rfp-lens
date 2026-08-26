from dataclasses import dataclass
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


EXPORT_HEADERS = [
    "요구사항",
    "분류",
    "필수 여부",
    "원문 근거",
    "원문 위치",
    "중요도",
    "제안서 반영 위치",
    "상태",
    "메모",
]


@dataclass(frozen=True)
class ExportRow:
    requirement: str
    category: str
    mandatory: str
    evidence_quote: str
    source_location: str
    importance: str
    proposal_section: str
    status: str
    owner_note: str


def _safe_cell(value: str) -> str:
    if value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def build_compliance_workbook(rows: list[ExportRow]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "컴플라이언스"
    sheet.append(EXPORT_HEADERS)
    for row in rows:
        sheet.append([_safe_cell(value) for value in row.__dict__.values()])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = [42, 20, 12, 48, 30, 14, 28, 18, 36]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:I{max(sheet.max_row, 1)}"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
