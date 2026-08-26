import logging
from io import BytesIO
from uuid import UUID, uuid4
from zipfile import ZIP_DEFLATED, ZIP_STORED, ZipFile

from fastapi.testclient import TestClient
from openpyxl import load_workbook
import pytest
from sqlalchemy import select
from sqlalchemy.orm import Session, sessionmaker

from app.core.celery_app import celery_app
from app.db.models import (
    AnalysisJob,
    Document,
    DocumentState,
    Evidence,
    JobState,
    Project,
    Requirement,
    RequirementCategory,
    ReviewState,
    User,
)
from app.documents.tasks import process_document


PASSWORD = "Correct-Horse-2026"
MAX_UPLOAD_BYTES = 26_214_400


def register_and_login(client: TestClient, email: str) -> dict[str, str]:
    assert (
        client.post(
            "/api/auth/register", json={"email": email, "password": PASSWORD}
        ).status_code
        == 201
    )
    token = client.post(
        "/api/auth/token", data={"username": email, "password": PASSWORD}
    ).json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


def create_project(client: TestClient, headers: dict[str, str]) -> str:
    response = client.post("/api/projects", headers=headers, json={"name": "보안 점검"})
    assert response.status_code == 201
    return response.json()["id"]


def build_zip(members: list[tuple[str, bytes]]) -> bytes:
    output = BytesIO()
    with ZipFile(output, "w") as archive:
        for name, data in members:
            stored = name == "mimetype"
            archive.writestr(
                name,
                data,
                compress_type=ZIP_STORED if stored else ZIP_DEFLATED,
            )
    return output.getvalue()


def valid_hwpx_bytes(section_xml: str | None = None) -> bytes:
    section = section_xml or (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<hs:sec xmlns:hs="http://www.owpml.org/owpml/2021/section"'
        '        xmlns:hp="http://www.owpml.org/owpml/2021/paragraph">'
        "<hp:p><hp:run><hp:t>중소기업만 신청 가능</hp:t></hp:run></hp:p>"
        "</hs:sec>"
    )
    content_hpf = (
        '<opf:package xmlns:opf="http://www.idpf.org/2007/opf">'
        '<opf:manifest><opf:item id="s0" href="section0.xml" /></opf:manifest>'
        '<opf:spine><opf:itemref idref="s0" /></opf:spine>'
        "</opf:package>"
    )
    return build_zip(
        [
            ("mimetype", b"application/hwp+zip"),
            ("Contents/content.hpf", content_hpf.encode()),
            ("Contents/section0.xml", section.encode()),
        ]
    )


def upload_document(
    client: TestClient,
    headers: dict[str, str],
    project_id: str,
    filename: str,
    payload: bytes,
):
    return client.post(
        f"/api/projects/{project_id}/documents",
        headers=headers,
        files={"file": (filename, payload, "application/octet-stream")},
    )


def test_oversized_upload_returns_file_too_large(client: TestClient) -> None:
    headers = register_and_login(client, "sec-oversize@example.com")
    project_id = create_project(client, headers)

    response = upload_document(
        client,
        headers,
        project_id,
        "big.pdf",
        b"%PDF-1.7\n" + b"a" * MAX_UPLOAD_BYTES,
    )

    assert response.status_code == 413
    assert response.json()["detail"]["code"] == "file_too_large"


def test_forged_hwpx_extension_with_wrong_mimetype_is_rejected(
    client: TestClient,
) -> None:
    headers = register_and_login(client, "sec-forged@example.com")
    project_id = create_project(client, headers)
    forged = build_zip(
        [("mimetype", b"application/zip"), ("Contents/content.hpf", b"<package />")]
    )

    response = upload_document(client, headers, project_id, "notice.hwpx", forged)

    assert response.status_code == 422
    assert response.json()["detail"]["code"] == "invalid_hwpx_mimetype"


def test_legacy_hwp_binary_is_rejected(client: TestClient) -> None:
    headers = register_and_login(client, "sec-legacy@example.com")
    project_id = create_project(client, headers)
    ole_hwp = bytes.fromhex("D0CF11E0A1B11AE1") + b"\x00" * 64

    response = upload_document(client, headers, project_id, "notice.hwp", ole_hwp)

    assert response.status_code == 422
    assert response.json()["detail"]["code"] == "legacy_hwp_unsupported"


def test_zip_traversal_member_is_rejected(client: TestClient) -> None:
    headers = register_and_login(client, "sec-traversal@example.com")
    project_id = create_project(client, headers)
    malicious = build_zip(
        [
            ("mimetype", b"application/hwp+zip"),
            ("Contents/content.hpf", b"<package />"),
            ("../escape.xml", b"<evil />"),
        ]
    )

    response = upload_document(client, headers, project_id, "notice.hwpx", malicious)

    assert response.status_code == 422
    detail = response.json()["detail"]
    assert detail["code"] == "invalid_hwpx"
    assert detail["message"] == "unsafe_member_path"


def test_excessive_zip_entries_are_rejected(client: TestClient) -> None:
    headers = register_and_login(client, "sec-members@example.com")
    project_id = create_project(client, headers)
    members: list[tuple[str, bytes]] = [
        ("mimetype", b"application/hwp+zip"),
        ("Contents/content.hpf", b"<package />"),
    ]
    members.extend((f"Contents/filler{index}.bin", b"x") for index in range(499))
    assert len(members) > 500

    response = upload_document(
        client, headers, project_id, "notice.hwpx", build_zip(members)
    )

    assert response.status_code == 422
    detail = response.json()["detail"]
    assert detail["code"] == "invalid_hwpx"
    assert detail["message"] == "too_many_members"


def test_compression_ratio_bomb_is_rejected(client: TestClient) -> None:
    headers = register_and_login(client, "sec-ratio@example.com")
    project_id = create_project(client, headers)
    bomb = build_zip(
        [
            ("mimetype", b"application/hwp+zip"),
            ("Contents/content.hpf", b"<package />"),
            ("Contents/payload.bin", b"0" * (5 * 1024 * 1024)),
        ]
    )

    response = upload_document(client, headers, project_id, "notice.hwpx", bomb)

    assert response.status_code == 422
    detail = response.json()["detail"]
    assert detail["code"] == "invalid_hwpx"
    assert detail["message"] == "compression_ratio_exceeded"


@pytest.fixture
def eager_processing(
    monkeypatch: pytest.MonkeyPatch, db_session: Session, client: TestClient
):
    import app.documents.tasks as document_tasks

    factory = sessionmaker(
        bind=db_session.connection(), autoflush=False, expire_on_commit=False
    )
    monkeypatch.setattr(document_tasks, "task_session_factory", factory)
    monkeypatch.setattr(document_tasks, "task_settings", client.app.state.settings)

    previous_eager = celery_app.conf.task_always_eager
    previous_propagates = celery_app.conf.task_eager_propagates
    celery_app.conf.task_always_eager = True
    celery_app.conf.task_eager_propagates = True
    yield
    celery_app.conf.task_always_eager = previous_eager
    celery_app.conf.task_eager_propagates = previous_propagates


def test_xml_entity_hwpx_fails_processing_without_requirements(
    client: TestClient, db_session: Session, eager_processing
) -> None:
    headers = register_and_login(client, "sec-entity@example.com")
    project_id = create_project(client, headers)
    entity_section = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        "<!DOCTYPE sec [<!ENTITY secret SYSTEM \"file:///C:/Windows/win.ini\">]>"
        '<hs:sec xmlns:hs="http://www.owpml.org/owpml/2021/section"'
        '        xmlns:hp="http://www.owpml.org/owpml/2021/paragraph">'
        "<hp:p><hp:run><hp:t>&secret;</hp:t></hp:run></hp:p>"
        "</hs:sec>"
    )
    uploaded = upload_document(
        client, headers, project_id, "notice.hwpx", valid_hwpx_bytes(entity_section)
    )
    assert uploaded.status_code == 201
    document_id = uploaded.json()["id"]

    result = process_document.run(document_id)
    db_session.expire_all()
    refreshed = db_session.get(Document, UUID(document_id))
    requirements = db_session.scalars(select(Requirement)).all()
    jobs = db_session.scalars(select(AnalysisJob)).all()

    assert result == DocumentState.FAILED.value
    assert refreshed.state == DocumentState.FAILED
    assert refreshed.error_code == "unsafe_xml"
    assert requirements == []
    assert jobs == []


def seed_requirement(db: Session, *, owner_email: str) -> Requirement:
    user = db.scalar(select(User).where(User.email == owner_email))
    assert user is not None
    project = Project(owner=user, name="보안 점검 프로젝트")
    document = Document(
        project=project,
        original_name="rfp.hwpx",
        media_type="application/hwp+zip",
        checksum_sha256="a" * 64,
        storage_key=f"fixture/{uuid4()}",
        state=DocumentState.REVIEW_REQUIRED,
    )
    job = AnalysisJob(
        project=project,
        document=document,
        state=JobState.SUCCEEDED,
        provider="fake",
        model="fixture",
        prompt_version="requirements-v1",
    )
    requirement = Requirement(
        project=project,
        document=document,
        analysis_job=job,
        text="중소기업만 신청 가능",
        category=RequirementCategory.ELIGIBILITY,
        mandatory=True,
        confidence="high",
        review_state=ReviewState.PENDING,
    )
    db.add(requirement)
    db.commit()
    db.refresh(requirement)
    return requirement


def test_cross_user_requirement_patch_and_export_return_not_found(
    client: TestClient, db_session: Session
) -> None:
    owner_email = "sec-owner@example.com"
    owner = register_and_login(client, owner_email)
    outsider = register_and_login(client, "sec-outsider@example.com")
    requirement = seed_requirement(db_session, owner_email=owner_email)
    project_id = str(requirement.project_id)

    listed = client.get(f"/api/projects/{project_id}/documents", headers=outsider)
    patched = client.patch(
        f"/api/projects/{project_id}/requirements/{requirement.id}",
        headers=outsider,
        json={
            "review_state": "confirmed",
            "updated_at": requirement.updated_at.isoformat(),
        },
    )
    exported = client.get(
        f"/api/projects/{project_id}/compliance.xlsx", headers=outsider
    )

    assert listed.status_code == 404
    assert patched.status_code == 404
    assert exported.status_code == 404
    assert client.get(
        f"/api/projects/{project_id}/compliance", headers=owner
    ).status_code in {200, 404}


def test_formula_injection_values_are_escaped_in_export(
    client: TestClient, db_session: Session
) -> None:
    owner_email = "sec-formula@example.com"
    headers = register_and_login(client, owner_email)
    requirement = seed_requirement(db_session, owner_email=owner_email)
    project_id = str(requirement.project_id)
    confirmed = client.patch(
        f"/api/projects/{project_id}/requirements/{requirement.id}",
        headers=headers,
        json={
            "review_state": "confirmed",
            "updated_at": requirement.updated_at.isoformat(),
        },
    )
    assert confirmed.status_code == 200
    item = client.get(f"/api/projects/{project_id}/compliance", headers=headers).json()[0]
    injected = '=HYPERLINK("https://evil.example","눌러주세요")'
    saved = client.patch(
        f"/api/projects/{project_id}/compliance/{item['id']}",
        headers=headers,
        json={"updated_at": item["updated_at"], "proposal_section": injected},
    )
    assert saved.status_code == 200

    exported = client.get(f"/api/projects/{project_id}/compliance.xlsx", headers=headers)
    sheet = load_workbook(BytesIO(exported.content)).active
    proposal_cell = sheet.cell(row=2, column=7).value

    assert exported.status_code == 200
    assert isinstance(proposal_cell, str)
    assert proposal_cell.startswith("'=")


def test_logs_do_not_contain_secrets_or_document_text(
    client: TestClient, caplog: pytest.LogCaptureFixture
) -> None:
    password_marker = "Sec-ret-Passphrase-2026!"
    document_marker = "극비-공고문-전문-마커"
    wrong_password = "totally-wrong-password"
    outsider = register_and_login(client, "sec-log-outsider@example.com")
    with caplog.at_level(logging.DEBUG):
        client.post(
            "/api/auth/register",
            json={"email": "sec-logs@example.com", "password": password_marker},
        )
        failed_login = client.post(
            "/api/auth/token",
            data={"username": "sec-logs@example.com", "password": wrong_password},
        )
        assert failed_login.status_code == 401
        token_response = client.post(
            "/api/auth/token",
            data={"username": "sec-logs@example.com", "password": password_marker},
        )
        token = token_response.json()["access_token"]
        headers = {"Authorization": f"Bearer {token}"}
        project_id = create_project(client, headers)
        upload_document(
            client,
            headers,
            project_id,
            "secret.pdf",
            b"%PDF-1.7\n" + document_marker.encode(),
        )
        upload_document(client, headers, project_id, "notice.txt", b"not a document")
        client.get(f"/api/projects/{project_id}", headers=outsider)

    captured = "\n".join(record.getMessage() for record in caplog.records)
    assert password_marker not in captured
    assert wrong_password not in captured
    assert token not in captured
    assert document_marker not in captured
    assert PASSWORD not in captured
