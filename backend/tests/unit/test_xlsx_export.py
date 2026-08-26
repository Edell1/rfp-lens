from io import BytesIO

from openpyxl import load_workbook

from app.compliance.export import EXPORT_HEADERS, ExportRow, build_compliance_workbook


def test_xlsx_headers_and_row_values_are_exact() -> None:
    content = build_compliance_workbook(
        [
            ExportRow(
                requirement="중소기업만 신청 가능",
                category="eligibility",
                mandatory="필수",
                evidence_quote="중소기업만 신청 가능",
                source_location="section0.xml 문단 1",
                importance="high",
                proposal_section="2. 연구개발 필요성",
                status="in_progress",
                owner_note="확인 완료",
            )
        ]
    )
    sheet = load_workbook(BytesIO(content)).active

    assert sheet.title == "컴플라이언스"
    assert [cell.value for cell in sheet[1]] == EXPORT_HEADERS
    assert [cell.value for cell in sheet[2]] == [
        "중소기업만 신청 가능",
        "eligibility",
        "필수",
        "중소기업만 신청 가능",
        "section0.xml 문단 1",
        "high",
        "2. 연구개발 필요성",
        "in_progress",
        "확인 완료",
    ]
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:I2"


def test_xlsx_escapes_formula_like_values() -> None:
    content = build_compliance_workbook(
        [
            ExportRow(
                requirement="=HYPERLINK(\"bad\")",
                category="other",
                mandatory="필수",
                evidence_quote="+SUM(1,1)",
                source_location="-1",
                importance="required",
                proposal_section="@hidden",
                status="not_started",
                owner_note="safe",
            )
        ]
    )
    row = [cell.value for cell in load_workbook(BytesIO(content)).active[2]]

    assert row[0].startswith("'=")
    assert row[3].startswith("'+")
    assert row[4].startswith("'-")
    assert row[6].startswith("'@")
